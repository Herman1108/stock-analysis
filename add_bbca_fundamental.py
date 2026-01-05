"""
Script to add Fundamental data to BBCA Excel file
Starting from column AL with attractive, colorful design
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load workbook
wb = openpyxl.load_workbook('C:/doc Herman/bbca.xlsx')
ws = wb.active

# Define styles - Premium blue theme for banking
header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')  # Deep blue
section_green = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')  # Deep green
section_orange = PatternFill(start_color='E65100', end_color='E65100', fill_type='solid')  # Deep orange
section_purple = PatternFill(start_color='4A148C', end_color='4A148C', fill_type='solid')  # Deep purple
section_teal = PatternFill(start_color='006064', end_color='006064', fill_type='solid')  # Deep teal
section_pink = PatternFill(start_color='880E4F', end_color='880E4F', fill_type='solid')  # Deep pink
section_navy = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')  # Deep indigo
section_gold = PatternFill(start_color='FF8F00', end_color='FF8F00', fill_type='solid')  # Amber

data_mint = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')  # Light green
data_peach = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')  # Light orange
data_lavender = PatternFill(start_color='EDE7F6', end_color='EDE7F6', fill_type='solid')  # Light purple
data_sky = PatternFill(start_color='E0F7FA', end_color='E0F7FA', fill_type='solid')  # Light cyan
data_cream = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')  # Light amber
data_rose = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')  # Light pink
data_blue = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')  # Light blue

header_font = Font(bold=True, color='FFFFFF', size=12)
section_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, color='263238', size=10)
normal_font = Font(color='263238', size=10)
italic_font = Font(italic=True, color='607D8B', size=9)
number_font = Font(color='0D47A1', size=10)

thin_border = Border(
    left=Side(style='thin', color='90A4AE'),
    right=Side(style='thin', color='90A4AE'),
    top=Side(style='thin', color='90A4AE'),
    bottom=Side(style='thin', color='90A4AE')
)

# Starting column AL = 38
start_col = 38
row = 1

# === MAIN HEADER ===
cell = ws.cell(row=row, column=start_col, value='ANALISIS FUNDAMENTAL')
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal='center')
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row = 3
# === MARKET OVERVIEW ===
cell = ws.cell(row=row, column=start_col, value='GAMBARAN UMUM')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Per 30 September 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
market_data = [
    ('Tahun Fiskal', 'Desember'),
    ('Total Saham Beredar', '123,28 Miliar lembar'),
    ('Kapitalisasi Pasar', 'Rp 989,28 Triliun'),
    ('Indeks Saham', '22.915,5'),
]

for label, value in market_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_mint
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_mint
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === FINANCIAL POSITION ===
cell = ws.cell(row=row, column=start_col, value='POSISI KEUANGAN')
cell.font = section_font
cell.fill = section_orange
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
financial_data = [
    ('Pendapatan Usaha', 'Rp 95,92 Triliun'),
    ('Total Aset', 'Rp 1.538,50 Triliun'),
    ('Total Liabilitas', 'Rp 1.261,87 Triliun'),
    ('Total Ekuitas', 'Rp 276,42 Triliun'),
    ('Belanja Modal (CapEx)', 'Rp 1,10 Triliun'),
    ('Biaya Operasional', 'Rp 32,11 Triliun'),
    ('Arus Kas Operasional', 'Rp 65,93 Triliun'),
    ('Arus Kas Bersih', 'Rp 14,78 Triliun'),
    ('Laba Usaha', 'Rp 53,77 Triliun'),
    ('Laba Tahun Berjalan', 'Rp 43,40 Triliun'),
]

for label, value in financial_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_peach
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_peach
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === PER SHARE DATA ===
cell = ws.cell(row=row, column=start_col, value='DATA PER LEMBAR SAHAM')
cell.font = section_font
cell.fill = section_purple
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
per_share_data = [
    ('Dividen Per Saham (DPS)', 'Rp 55'),
    ('Laba Per Saham (EPS)', 'Rp 469,38'),
    ('Pendapatan Per Saham (RPS)', 'Rp 1.037,42'),
    ('Nilai Buku Per Saham (BVPS)', 'Rp 2.242,27'),
    ('Arus Kas Per Saham (CFPS)', 'Rp 713,12'),
    ('Kas Setara Per Saham (CEPS)', 'Rp 727,19'),
    ('Aset Bersih Per Saham (NAVS)', 'Rp 2.244,05'),
]

for label, value in per_share_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_lavender
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_lavender
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === VALUATION METRICS ===
cell = ws.cell(row=row, column=start_col, value='METRIK VALUASI')
cell.font = section_font
cell.fill = section_teal
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
valuation_data = [
    ('Yield Dividen', '0,69%'),
    ('Rasio Harga/Laba (PER)', '17,10x'),
    ('Rasio Harga/Pendapatan (PSR)', '7,74x'),
    ('Rasio Harga/Nilai Buku (PBV)', '3,58x'),
    ('Rasio Harga/Arus Kas (PCFR)', '11,25x'),
]

for label, value in valuation_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_sky
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_sky
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === PROFITABILITY ===
cell = ws.cell(row=row, column=start_col, value='RASIO PROFITABILITAS')
cell.font = section_font
cell.fill = section_pink
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
profitability_data = [
    ('Rasio Pembayaran Dividen (DPR)', '11,72%'),
    ('Marjin Laba Usaha (OPM)', '56,06%'),
    ('Marjin Laba Bersih (NPM)', '45,25%'),
    ('Imbal Hasil Ekuitas (ROE)', '20,93%'),
    ('Imbal Hasil Aset (ROA)', '3,76%'),
]

for label, value in profitability_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_rose
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_rose
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === LIQUIDITY ===
cell = ws.cell(row=row, column=start_col, value='RASIO LIKUIDITAS & LEVERAGE')
cell.font = section_font
cell.fill = section_navy
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
liquidity_data = [
    ('Rasio Utang/Ekuitas (DER)', '456,51%'),
    ('Rasio Kas (Cash Ratio)', '7,11%'),
    ('Rasio Cepat (Quick Ratio)', '99,69%'),
    ('Rasio Lancar (Current Ratio)', '99,69%'),
]

for label, value in liquidity_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_cream
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_cream
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === DIVIDEND HISTORY ===
cell = ws.cell(row=row, column=start_col, value='RIWAYAT DIVIDEN')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
dividend_data = [
    ('Dividen Interim 2025', 'Rp 55 (Ex: 03/12/2025)'),
    ('Dividen Final 2024', 'Rp 250 (Ex: 21/03/2025)'),
    ('Dividen Interim 2024', 'Rp 50 (Ex: 21/11/2024)'),
    ('Dividen Final 2023', 'Rp 227,5 (Ex: 25/03/2024)'),
    ('Dividen Interim 2023', 'Rp 42,5 (Ex: 04/12/2023)'),
    ('Dividen Final 2022', 'Rp 170 (Ex: 29/03/2023)'),
]

for label, value in dividend_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_mint
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_mint
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === QUARTERLY PERFORMANCE ===
cell = ws.cell(row=row, column=start_col, value='KINERJA KUARTALAN 2025')
cell.font = section_font
cell.fill = section_orange
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
quarterly_data = [
    ('Pendapatan Q1 2025', 'Rp 31,4 Triliun'),
    ('Pendapatan Q2 2025', 'Rp 32,1 Triliun'),
    ('Pendapatan Q3 2025', 'Rp 32,5 Triliun'),
    ('Total Pendapatan 9M 2025', 'Rp 96 Triliun'),
    ('Laba Bersih Q1 2025', 'Rp 14,15 Triliun'),
    ('Laba Bersih Q2 2025', 'Rp 14,87 Triliun'),
    ('Laba Bersih Q3 2025', 'Rp 14,38 Triliun'),
    ('Total Laba Bersih 9M 2025', 'Rp 43,4 Triliun'),
    ('EPS Kumulatif 2025', 'Rp 469'),
]

for label, value in quarterly_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_peach
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_peach
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === HISTORICAL COMPARISON ===
cell = ws.cell(row=row, column=start_col, value='PERBANDINGAN TAHUNAN')
cell.font = section_font
cell.fill = section_gold
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
historical_data = [
    ('EPS 2022', 'Rp 330'),
    ('EPS 2023', 'Rp 395'),
    ('EPS 2024', 'Rp 445'),
    ('EPS 2025 (9M)', 'Rp 469'),
    ('DPS 2022', 'Rp 205'),
    ('DPS 2023', 'Rp 270'),
    ('DPS 2024', 'Rp 300'),
    ('Pendapatan 2022', 'Rp 96 Triliun'),
    ('Pendapatan 2023', 'Rp 112 Triliun'),
    ('Pendapatan 2024', 'Rp 121 Triliun'),
    ('Laba Bersih 2022', 'Rp 41 Triliun'),
    ('Laba Bersih 2023', 'Rp 49 Triliun'),
    ('Laba Bersih 2024', 'Rp 55 Triliun'),
]

for label, value in historical_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_blue
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_blue
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

# Set column widths
ws.column_dimensions[get_column_letter(start_col)].width = 38
ws.column_dimensions[get_column_letter(start_col+1)].width = 32

# Save workbook
wb.save('C:/doc Herman/bbca.xlsx')
print('BBCA Fundamental data added to Excel successfully!')
print(f'Data written from column AL (38) to AM (39)')
print(f'Total rows used: {row}')
