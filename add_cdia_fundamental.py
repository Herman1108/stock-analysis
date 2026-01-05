"""
Script to add Fundamental data to CDIA Excel file
Starting from column AL with attractive, colorful design
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load workbook
wb = openpyxl.load_workbook('C:/doc Herman/cdia.xlsx')
ws = wb.active

# Define styles - Soft and cheerful colors
header_fill = PatternFill(start_color='1E88E5', end_color='1E88E5', fill_type='solid')  # Bright blue
section_green = PatternFill(start_color='43A047', end_color='43A047', fill_type='solid')  # Fresh green
section_orange = PatternFill(start_color='FB8C00', end_color='FB8C00', fill_type='solid')  # Warm orange
section_purple = PatternFill(start_color='8E24AA', end_color='8E24AA', fill_type='solid')  # Rich purple
section_teal = PatternFill(start_color='00ACC1', end_color='00ACC1', fill_type='solid')  # Teal cyan
section_pink = PatternFill(start_color='D81B60', end_color='D81B60', fill_type='solid')  # Rose pink
section_navy = PatternFill(start_color='3949AB', end_color='3949AB', fill_type='solid')  # Indigo

data_mint = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')  # Soft mint
data_peach = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')  # Soft peach
data_lavender = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')  # Soft lavender
data_sky = PatternFill(start_color='E1F5FE', end_color='E1F5FE', fill_type='solid')  # Soft sky
data_cream = PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid')  # Soft cream
data_rose = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')  # Soft rose

header_font = Font(bold=True, color='FFFFFF', size=12)
section_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, color='37474F', size=10)
normal_font = Font(color='37474F', size=10)
italic_font = Font(italic=True, color='607D8B', size=9)
number_font = Font(color='1565C0', size=10)

thin_border = Border(
    left=Side(style='thin', color='B0BEC5'),
    right=Side(style='thin', color='B0BEC5'),
    top=Side(style='thin', color='B0BEC5'),
    bottom=Side(style='thin', color='B0BEC5')
)

# Starting column AL = 38
start_col = 38
row = 1

# === MAIN HEADER ===
cell = ws.cell(row=row, column=start_col, value='ANALISIS FUNDAMENTAL')
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal='center')
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row = 3
# === MARKET OVERVIEW ===
cell = ws.cell(row=row, column=start_col, value='GAMBARAN UMUM PASAR')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Per 30 September 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
market_data = [
    ('Periode Tutup Buku', 'Desember'),
    ('Jumlah Saham Beredar', '124,83 Miliar lembar'),
    ('Nilai Kapitalisasi Pasar', 'Rp 214,08 Triliun'),
    ('Indeks Saham', '902,6'),
]

for label, value in market_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_mint
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_mint
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === FINANCIAL POSITION ===
cell = ws.cell(row=row, column=start_col, value='KONDISI KEUANGAN')
cell.font = section_font
cell.fill = section_orange
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
financial_data = [
    ('Pendapatan Usaha', 'Rp 1,74 Triliun'),
    ('Total Aset', 'Rp 26,58 Triliun'),
    ('Total Liabilitas', 'Rp 7,47 Triliun'),
    ('Total Ekuitas', 'Rp 17,16 Triliun'),
    ('Belanja Modal (CapEx)', 'Rp 1,88 Triliun'),
    ('Beban Operasional', 'Rp 146,79 Miliar'),
    ('Arus Kas dari Operasi', 'Rp 128,15 Miliar'),
    ('Arus Kas Bersih', 'Rp 4,92 Triliun'),
    ('Laba Usaha', 'Rp 252,60 Miliar'),
    ('Laba Tahun Berjalan', 'Rp 1,29 Triliun'),
]

for label, value in financial_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_peach
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_peach
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === PER SHARE DATA ===
cell = ws.cell(row=row, column=start_col, value='NILAI PER LEMBAR SAHAM')
cell.font = section_font
cell.fill = section_purple
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
per_share_data = [
    ('Dividen Per Saham (DPS)', '-'),
    ('Laba Per Saham (EPS)', 'Rp 13,79'),
    ('Pendapatan Per Saham (RPS)', 'Rp 18,61'),
    ('Nilai Buku Per Saham (BVPS)', 'Rp 137,38'),
    ('Arus Kas Per Saham (CFPS)', 'Rp 1,33'),
    ('Kas Setara Per Saham (CEPS)', 'Rp 63,46'),
    ('Aset Bersih Per Saham (NAVS)', 'Rp 153,16'),
]

for label, value in per_share_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_lavender
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_lavender
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === VALUATION METRICS ===
cell = ws.cell(row=row, column=start_col, value='INDIKATOR VALUASI')
cell.font = section_font
cell.fill = section_teal
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
valuation_data = [
    ('Imbal Hasil Dividen', '-'),
    ('Rasio Harga terhadap Laba (PER)', '124,38x'),
    ('Rasio Harga terhadap Penjualan (PSR)', '92,18x'),
    ('Rasio Harga terhadap Nilai Buku (PBV)', '12,48x'),
    ('Rasio Harga terhadap Arus Kas (PCFR)', '1.290,48x'),
]

for label, value in valuation_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_sky
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_sky
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === PROFITABILITY ===
cell = ws.cell(row=row, column=start_col, value='TINGKAT PROFITABILITAS')
cell.font = section_font
cell.fill = section_pink
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
profitability_data = [
    ('Rasio Pembayaran Dividen (DPR)', '-'),
    ('Marjin Laba Kotor (GPM)', '22,94%'),
    ('Marjin Laba Usaha (OPM)', '14,51%'),
    ('Marjin Laba Bersih (NPM)', '74,04%'),
    ('Marjin EBIT (EBITM)', '100,81%'),
    ('Tingkat Pengembalian Ekuitas (ROE)', '10,01%'),
    ('Tingkat Pengembalian Aset (ROA)', '6,47%'),
]

for label, value in profitability_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_rose
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_rose
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === LIQUIDITY ===
cell = ws.cell(row=row, column=start_col, value='RASIO LIKUIDITAS & SOLVABILITAS')
cell.font = section_font
cell.fill = section_navy
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
liquidity_data = [
    ('Rasio Utang terhadap Ekuitas (DER)', '43,53%'),
    ('Rasio Kas (Cash Ratio)', '1.075,81%'),
    ('Rasio Cepat (Quick Ratio)', '1.368,16%'),
    ('Rasio Lancar (Current Ratio)', '1.370,59%'),
]

for label, value in liquidity_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_cream
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_cream
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === DIVIDEND HISTORY ===
cell = ws.cell(row=row, column=start_col, value='JADWAL DIVIDEN')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
dividend_data = [
    ('Dividen 2025', 'Ex: 12/01/2026, Bayar: 29/01/2026'),
]

for label, value in dividend_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_mint
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_mint
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === QUARTERLY PERFORMANCE ===
cell = ws.cell(row=row, column=start_col, value='KINERJA KUARTALAN 2025')
cell.font = section_font
cell.fill = section_orange
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
quarterly_data = [
    ('Pendapatan Q2 2025', 'Rp 1,09 Triliun'),
    ('Pendapatan Q3 2025', 'Rp 656 Miliar'),
    ('Laba Bersih Q2 2025', 'Rp 1,10 Triliun'),
    ('Laba Bersih Q3 2025', 'Rp 188 Miliar'),
    ('Total Pendapatan 9M 2025', 'Rp 1,74 Triliun'),
    ('Total Laba Bersih 9M 2025', 'Rp 1,72 Triliun'),
    ('EPS Kumulatif 2025', 'Rp 14'),
]

for label, value in quarterly_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_peach
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_peach
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === HISTORICAL COMPARISON ===
cell = ws.cell(row=row, column=start_col, value='PERBANDINGAN HISTORIS')
cell.font = section_font
cell.fill = section_purple
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
historical_data = [
    ('EPS 2023', 'Rp 0'),
    ('EPS 2024', 'Rp 4'),
    ('EPS 2025 (9M)', 'Rp 14'),
    ('Pendapatan 2023', 'Rp 1,17 Triliun'),
    ('Pendapatan 2024', 'Rp 1,65 Triliun'),
    ('Laba Bersih 2023', 'Rp 3 Miliar'),
    ('Laba Bersih 2024', 'Rp 495 Miliar'),
]

for label, value in historical_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_lavender
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_lavender
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

# Set column widths
ws.column_dimensions[get_column_letter(start_col)].width = 40
ws.column_dimensions[get_column_letter(start_col+1)].width = 32

# Save workbook
wb.save('C:/doc Herman/cdia.xlsx')
print('CDIA Fundamental data added to Excel successfully!')
print(f'Data written from column AL (38) to AM (39)')
print(f'Total rows used: {row}')
