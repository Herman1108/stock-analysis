"""
Script to add Company Profile data to BBCA Excel file
Starting from column AI with attractive, colorful design
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load workbook
wb = openpyxl.load_workbook('C:/doc Herman/bbca.xlsx')
ws = wb.active

# Define styles
header_fill = PatternFill(start_color='4A90D9', end_color='4A90D9', fill_type='solid')
section_green = PatternFill(start_color='7BC97F', end_color='7BC97F', fill_type='solid')
highlight_fill = PatternFill(start_color='F9E79F', end_color='F9E79F', fill_type='solid')
data_fill = PatternFill(start_color='E8F6F3', end_color='E8F6F3', fill_type='solid')
pink_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
lavender_fill = PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid')
orange_fill = PatternFill(start_color='E59866', end_color='E59866', fill_type='solid')
sky_blue = PatternFill(start_color='5DADE2', end_color='5DADE2', fill_type='solid')
purple_fill = PatternFill(start_color='AF7AC5', end_color='AF7AC5', fill_type='solid')
teal_fill = PatternFill(start_color='48C9B0', end_color='48C9B0', fill_type='solid')
steel_blue = PatternFill(start_color='5499C7', end_color='5499C7', fill_type='solid')
coral_fill = PatternFill(start_color='EC7063', end_color='EC7063', fill_type='solid')

header_font = Font(bold=True, color='FFFFFF', size=12)
section_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, color='2C3E50', size=10)
normal_font = Font(color='2C3E50', size=10)
italic_font = Font(italic=True, color='566573', size=9)

thin_border = Border(
    left=Side(style='thin', color='BDC3C7'),
    right=Side(style='thin', color='BDC3C7'),
    top=Side(style='thin', color='BDC3C7'),
    bottom=Side(style='thin', color='BDC3C7')
)

start_col = 35
row = 1

# === MAIN HEADER ===
cell = ws.cell(row=row, column=start_col, value='COMPANY PROFILE')
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal='center')
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row = 3
# === COMPANY IDENTITY ===
cell = ws.cell(row=row, column=start_col, value='IDENTITAS PERUSAHAAN')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
profile_data = [
    ('Nama Emiten', 'PT Bank Central Asia Tbk'),
    ('Kode Saham', 'BBCA'),
    ('Papan Pencatatan', 'Papan Utama (Main Board)'),
    ('Sektor Usaha', 'Keuangan (Financials)'),
    ('Sub Sektor', 'Perbankan'),
    ('Bidang Industri', 'Bank Umum'),
    ('Aktivitas Bisnis', 'Jasa perbankan dan layanan keuangan terintegrasi'),
]

for label, value in profile_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = data_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === COMPANY HISTORY ===
cell = ws.cell(row=row, column=start_col, value='SEJARAH & PENCATATAN')
cell.font = section_font
cell.fill = orange_fill
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
history_data = [
    ('Tanggal Pencatatan', '31 Mei 2000'),
    ('Tanggal Efektif', '11 Mei 2000'),
    ('Nilai Nominal', 'Rp 13 (setelah stock split)'),
    ('Harga Penawaran Perdana', 'Rp 1.400'),
    ('Jumlah Saham IPO', '662.400.000 lembar'),
    ('Dana IPO Terhimpun', 'Rp 927,36 Miliar'),
    ('Penjamin Emisi', 'PT Danareksa Sekuritas, PT Bahana Securities'),
    ('Biro Administrasi Efek', 'PT Raya Saham Registra'),
]

for label, value in history_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = pink_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = pink_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === COMPANY BACKGROUND ===
cell = ws.cell(row=row, column=start_col, value='LATAR BELAKANG PERUSAHAAN')
cell.font = section_font
cell.fill = sky_blue
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
background_text = [
    'Bank swasta terbesar di Indonesia dengan layanan perbankan komprehensif.',
    'Memiliki anak usaha: BCA Finance, BCA Syariah, BCA Sekuritas, Asuransi BCA.',
    'Dikenal dengan jaringan ATM terluas dan layanan digital banking terdepan.',
    'Tahun 2017 mendirikan Central Capital Ventura untuk investasi fintech.',
    'Tercatat di BEI sejak Mei 2000, menjadi salah satu emiten blue chip unggulan.',
]

for text in background_text:
    ws.cell(row=row, column=start_col, value=text).font = normal_font
    ws.cell(row=row, column=start_col).fill = lavender_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)
    row += 1

row += 1
# === SHAREHOLDERS ===
cell = ws.cell(row=row, column=start_col, value='KOMPOSISI PEMEGANG SAHAM')
cell.font = section_font
cell.fill = purple_fill
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Per 30 November 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
shareholders = [
    ('PT Dwimuria Investama Andalan (Pengendali)', '54,94%'),
    ('Pihak Afiliasi Pengendali', '2,46%'),
    ('Robert Budi Hartono (Pengendali)', '0,02%'),
    ('Bambang Hartono (Pengendali)', '0,02%'),
    ('Publik (Non-Warkat)', '42,44%'),
    ('Saham Treasury', '0,05%'),
    ('Total Saham Beredar', '123.275.050.000 lembar'),
]

for label, value in shareholders:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = highlight_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = highlight_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === SHAREHOLDER COUNT HISTORY ===
cell = ws.cell(row=row, column=start_col, value='PERKEMBANGAN JUMLAH INVESTOR')
cell.font = section_font
cell.fill = teal_fill
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
investor_history = [
    ('November 2025', '569.882 (-16.061)'),
    ('Oktober 2025', '585.943 (-26.230)'),
    ('September 2025', '612.173 (+89.259)'),
    ('Agustus 2025', '522.914 (+13.514)'),
    ('Juli 2025', '509.400 (+27.678)'),
    ('Juni 2025', '481.722 (+22.761)'),
]

for label, value in investor_history:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = data_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === BOARD OF DIRECTORS ===
cell = ws.cell(row=row, column=start_col, value='JAJARAN DIREKSI')
cell.font = section_font
cell.fill = steel_blue
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Efektif 2 Juni 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
directors = [
    ('Presiden Direktur', 'Gregory Hendra Lembong'),
    ('Wakil Presiden Direktur', 'John Kosasih'),
    ('Wakil Presiden Direktur', 'Armand Wahyudi Hartono'),
    ('Direktur', 'Rudy Susanto'),
    ('Direktur', 'Hendra Tanumihardja'),
    ('Direktur', 'Antonius Widodo Mulyono'),
    ('Direktur', 'Frengky Chandra Kusuma'),
    ('Direktur', 'Haryanto Tiara Budiman'),
    ('Direktur', 'Vera Eve Lim'),
    ('Direktur', 'Santoso'),
    ('Direktur', 'Lianawaty Suwono'),
    ('Direktur', 'Tan Ho Hien (Subur Tan)'),
]

for label, value in directors:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = pink_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = pink_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === BOARD OF COMMISSIONERS ===
cell = ws.cell(row=row, column=start_col, value='DEWAN KOMISARIS')
cell.font = section_font
cell.fill = coral_fill
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Efektif 2 Juni 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
commissioners = [
    ('Presiden Komisaris', 'Jahja Setiaatmadja'),
    ('Komisaris', 'Tonny Kusnadi'),
    ('Komisaris Independen', 'Sumantri Slamet'),
    ('Komisaris Independen', 'Cyrillus Harinowo'),
    ('Komisaris Independen', 'Raden Pardede'),
]

for label, value in commissioners:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = lavender_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = lavender_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

# Set column widths
ws.column_dimensions[get_column_letter(start_col)].width = 38
ws.column_dimensions[get_column_letter(start_col+1)].width = 45

# Save workbook
wb.save('C:/doc Herman/bbca.xlsx')
print('BBCA Profile data added to Excel successfully!')
print(f'Data written from column AI (35) to AJ (36)')
print(f'Total rows used: {row}')
