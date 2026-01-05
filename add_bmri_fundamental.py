"""
Script to add Fundamental data to BMRI Excel file
Starting from column AL with attractive, colorful design (same format as CDIA)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load workbook
wb = openpyxl.load_workbook('C:/doc Herman/bmri.xlsx')
ws = wb.active

# Define styles - Same as CDIA format
header_fill = PatternFill(start_color='1E88E5', end_color='1E88E5', fill_type='solid')  # Bright blue
section_green = PatternFill(start_color='43A047', end_color='43A047', fill_type='solid')  # Fresh green
section_orange = PatternFill(start_color='FB8C00', end_color='FB8C00', fill_type='solid')  # Warm orange
section_purple = PatternFill(start_color='8E24AA', end_color='8E24AA', fill_type='solid')  # Rich purple
section_teal = PatternFill(start_color='00ACC1', end_color='00ACC1', fill_type='solid')  # Teal cyan
section_pink = PatternFill(start_color='D81B60', end_color='D81B60', fill_type='solid')  # Rose pink
section_navy = PatternFill(start_color='3949AB', end_color='3949AB', fill_type='solid')  # Indigo

data_mint = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')  # Soft mint
data_peach = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')  # Soft peach
data_lavender = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')  # Soft lavender
data_sky = PatternFill(start_color='E1F5FE', end_color='E1F5FE', fill_type='solid')  # Soft sky
data_cream = PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid')  # Soft cream
data_rose = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')  # Soft rose

header_font = Font(bold=True, color='FFFFFF', size=12)
section_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, color='37474F', size=10)
normal_font = Font(color='37474F', size=10)
italic_font = Font(italic=True, color='607D8B', size=9)
number_font = Font(color='1565C0', size=10)

thin_border = Border(
    left=Side(style='thin', color='B0BEC5'),
    right=Side(style='thin', color='B0BEC5'),
    top=Side(style='thin', color='B0BEC5'),
    bottom=Side(style='thin', color='B0BEC5')
)

# Starting column AL = 38
start_col = 38
row = 1

# === MAIN HEADER ===
cell = ws.cell(row=row, column=start_col, value='ANALISIS FUNDAMENTAL')
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal='center')
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row = 3
# === MARKET OVERVIEW ===
cell = ws.cell(row=row, column=start_col, value='GAMBARAN UMUM PASAR')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Per 30 September 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
market_data = [
    ('Periode Tutup Buku', 'Desember'),
    ('Jumlah Saham Beredar', '93,33 Miliar lembar'),
    ('Nilai Kapitalisasi Pasar', 'Rp 473,67 Triliun'),
    ('Indeks Saham', '1.529,2'),
]

for label, value in market_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_mint
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_mint
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === FINANCIAL POSITION ===
cell = ws.cell(row=row, column=start_col, value='KONDISI KEUANGAN')
cell.font = section_font
cell.fill = section_orange
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
financial_data = [
    ('Pendapatan Usaha', 'Rp 155,34 Triliun'),
    ('Total Aset', 'Rp 2.563,36 Triliun'),
    ('Total Liabilitas', 'Rp 2.249,52 Triliun'),
    ('Total Ekuitas', 'Rp 281,63 Triliun'),
    ('Belanja Modal (CapEx)', 'Rp 4,39 Triliun'),
    ('Beban Operasional', 'Rp 58,07 Triliun'),
    ('Arus Kas dari Operasi', 'Rp 57,99 Triliun'),
    ('Arus Kas Bersih', 'Rp -7,28 Triliun'),
    ('Laba Usaha', 'Rp 50,93 Triliun'),
    ('Laba Tahun Berjalan', 'Rp 37,73 Triliun'),
]

for label, value in financial_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_peach
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_peach
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === PER SHARE DATA ===
cell = ws.cell(row=row, column=start_col, value='NILAI PER LEMBAR SAHAM')
cell.font = section_font
cell.fill = section_purple
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
per_share_data = [
    ('Dividen Per Saham (DPS)', 'Rp 100'),
    ('Laba Per Saham (EPS)', 'Rp 539,00'),
    ('Pendapatan Per Saham (RPS)', 'Rp 2.219,11'),
    ('Nilai Buku Per Saham (BVPS)', 'Rp 3.017,48'),
    ('Arus Kas Per Saham (CFPS)', 'Rp 828,47'),
    ('Kas Setara Per Saham (CEPS)', 'Rp 2.321,73'),
    ('Aset Bersih Per Saham (NAVS)', 'Rp 3.362,55'),
]

for label, value in per_share_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_lavender
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_lavender
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === VALUATION METRICS ===
cell = ws.cell(row=row, column=start_col, value='INDIKATOR VALUASI')
cell.font = section_font
cell.fill = section_teal
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
valuation_data = [
    ('Imbal Hasil Dividen', '1,97%'),
    ('Rasio Harga terhadap Laba (PER)', '9,42x'),
    ('Rasio Harga terhadap Penjualan (PSR)', '2,29x'),
    ('Rasio Harga terhadap Nilai Buku (PBV)', '1,68x'),
    ('Rasio Harga terhadap Arus Kas (PCFR)', '6,13x'),
]

for label, value in valuation_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_sky
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_sky
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === PROFITABILITY ===
cell = ws.cell(row=row, column=start_col, value='TINGKAT PROFITABILITAS')
cell.font = section_font
cell.fill = section_pink
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
profitability_data = [
    ('Rasio Pembayaran Dividen (DPR)', '18,55%'),
    ('Marjin Laba Usaha (OPM)', '32,79%'),
    ('Marjin Laba Bersih (NPM)', '24,29%'),
    ('Tingkat Pengembalian Ekuitas (ROE)', '17,87%'),
    ('Tingkat Pengembalian Aset (ROA)', '1,96%'),
]

for label, value in profitability_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_rose
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_rose
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === LIQUIDITY ===
cell = ws.cell(row=row, column=start_col, value='RASIO LIKUIDITAS & SOLVABILITAS')
cell.font = section_font
cell.fill = section_navy
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
liquidity_data = [
    ('Rasio Utang terhadap Ekuitas (DER)', '798,75%'),
    ('Rasio Kas (Cash Ratio)', '10,27%'),
    ('Rasio Cepat (Quick Ratio)', '113,71%'),
    ('Rasio Lancar (Current Ratio)', '113,71%'),
]

for label, value in liquidity_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_cream
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_cream
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === DIVIDEND HISTORY ===
cell = ws.cell(row=row, column=start_col, value='RIWAYAT DIVIDEN')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
dividend_data = [
    ('Dividen 2025', 'Rp 100 (Ex: 06/01/2026)'),
    ('Dividen 2024', 'Rp 466,18 (Ex: 14/04/2025)'),
    ('Dividen 2023', 'Rp 353,96 (Ex: 20/03/2024)'),
    ('Dividen 2022', 'Rp 529,34 (Ex: 27/03/2023)'),
    ('Dividen 2021', 'Rp 360,64 (Ex: 21/03/2022)'),
]

for label, value in dividend_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_mint
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_mint
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === QUARTERLY PERFORMANCE ===
cell = ws.cell(row=row, column=start_col, value='KINERJA KUARTALAN 2025')
cell.font = section_font
cell.fill = section_orange
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
quarterly_data = [
    ('Pendapatan Q1 2025', 'Rp 50,8 Triliun'),
    ('Pendapatan Q2 2025', 'Rp 51,6 Triliun'),
    ('Pendapatan Q3 2025', 'Rp 53,0 Triliun'),
    ('Total Pendapatan 9M 2025', 'Rp 155,34 Triliun'),
    ('Laba Bersih Q1 2025', 'Rp 13,2 Triliun'),
    ('Laba Bersih Q2 2025', 'Rp 11,3 Triliun'),
    ('Laba Bersih Q3 2025', 'Rp 13,3 Triliun'),
    ('Total Laba Bersih 9M 2025', 'Rp 37,73 Triliun'),
    ('EPS Kumulatif 2025', 'Rp 539'),
]

for label, value in quarterly_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_peach
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_peach
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === HISTORICAL COMPARISON ===
cell = ws.cell(row=row, column=start_col, value='PERBANDINGAN HISTORIS')
cell.font = section_font
cell.fill = section_purple
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
historical_data = [
    ('EPS 2022', 'Rp 441'),
    ('EPS 2023', 'Rp 590'),
    ('EPS 2024', 'Rp 598'),
    ('EPS 2025 (9M)', 'Rp 539'),
    ('Pendapatan 2022', 'Rp 147 Triliun'),
    ('Pendapatan 2023', 'Rp 173 Triliun'),
    ('Pendapatan 2024', 'Rp 193 Triliun'),
    ('Laba Bersih 2022', 'Rp 41 Triliun'),
    ('Laba Bersih 2023', 'Rp 55 Triliun'),
    ('Laba Bersih 2024', 'Rp 56 Triliun'),
]

for label, value in historical_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_lavender
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_lavender
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

# Set column widths
ws.column_dimensions[get_column_letter(start_col)].width = 40
ws.column_dimensions[get_column_letter(start_col+1)].width = 32

# Save workbook
wb.save('C:/doc Herman/bmri.xlsx')
print('BMRI Fundamental data added to Excel successfully!')
print(f'Data written from column AL (38) to AM (39)')
print(f'Total rows used: {row}')
