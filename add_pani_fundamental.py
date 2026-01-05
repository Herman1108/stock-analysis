"""
Script to add Fundamental data to PANI Excel file
Starting from column AL with attractive, colorful design
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load workbook
wb = openpyxl.load_workbook('C:/doc Herman/pani.xlsx')
ws = wb.active

# Define styles - Soft and cheerful colors
header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')  # Deep blue
section_green = PatternFill(start_color='58B368', end_color='58B368', fill_type='solid')  # Fresh green
section_orange = PatternFill(start_color='F18F01', end_color='F18F01', fill_type='solid')  # Vibrant orange
section_purple = PatternFill(start_color='9B5DE5', end_color='9B5DE5', fill_type='solid')  # Soft purple
section_teal = PatternFill(start_color='00B4D8', end_color='00B4D8', fill_type='solid')  # Teal
section_pink = PatternFill(start_color='E56B6F', end_color='E56B6F', fill_type='solid')  # Soft coral
section_navy = PatternFill(start_color='355070', end_color='355070', fill_type='solid')  # Navy

data_mint = PatternFill(start_color='D8F3DC', end_color='D8F3DC', fill_type='solid')  # Light mint
data_peach = PatternFill(start_color='FFE5D9', end_color='FFE5D9', fill_type='solid')  # Light peach
data_lavender = PatternFill(start_color='E2D9F3', end_color='E2D9F3', fill_type='solid')  # Light lavender
data_sky = PatternFill(start_color='CAF0F8', end_color='CAF0F8', fill_type='solid')  # Light sky
data_cream = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')  # Cream
data_rose = PatternFill(start_color='FFE4E6', end_color='FFE4E6', fill_type='solid')  # Light rose

header_font = Font(bold=True, color='FFFFFF', size=12)
section_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, color='2C3E50', size=10)
normal_font = Font(color='2C3E50', size=10)
italic_font = Font(italic=True, color='566573', size=9)
number_font = Font(color='1A5276', size=10)

thin_border = Border(
    left=Side(style='thin', color='BDC3C7'),
    right=Side(style='thin', color='BDC3C7'),
    top=Side(style='thin', color='BDC3C7'),
    bottom=Side(style='thin', color='BDC3C7')
)

# Starting column AL = 38
start_col = 38
row = 1

# === MAIN HEADER ===
cell = ws.cell(row=row, column=start_col, value='FUNDAMENTAL ANALYSIS')
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal='center')
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row = 3
# === MARKET OVERVIEW ===
cell = ws.cell(row=row, column=start_col, value='RINGKASAN PASAR')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Data per 30 September 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
market_data = [
    ('Tahun Buku Berakhir', 'Desember'),
    ('Saham Beredar', '18,12 Miliar lembar'),
    ('Kapitalisasi Pasar', 'Rp 228,73 Triliun'),
    ('Indeks Saham', '165.248,7'),
]

for label, value in market_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_mint
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_mint
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === FINANCIAL POSITION ===
cell = ws.cell(row=row, column=start_col, value='POSISI KEUANGAN')
cell.font = section_font
cell.fill = section_orange
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
financial_data = [
    ('Total Pendapatan', 'Rp 3,10 Triliun'),
    ('Total Aset', 'Rp 49,46 Triliun'),
    ('Total Kewajiban', 'Rp 18,75 Triliun'),
    ('Total Ekuitas', 'Rp 22,49 Triliun'),
    ('Belanja Modal (CapEx)', 'Rp 1,17 Triliun'),
    ('Biaya Operasional', 'Rp 303,38 Miliar'),
    ('Arus Kas Operasi', 'Rp 214,22 Miliar'),
    ('Arus Kas Bersih', 'Rp 314,52 Miliar'),
    ('Laba Operasi', 'Rp 1,73 Triliun'),
    ('Laba Bersih', 'Rp 791,30 Miliar'),
]

for label, value in financial_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_peach
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_peach
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === PER SHARE DATA ===
cell = ws.cell(row=row, column=start_col, value='DATA PER LEMBAR SAHAM')
cell.font = section_font
cell.fill = section_purple
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
per_share_data = [
    ('Dividen Per Saham (DPS)', 'Rp 4'),
    ('Laba Per Saham (EPS)', 'Rp 58,24'),
    ('Pendapatan Per Saham (RPS)', 'Rp 228,06'),
    ('Nilai Buku Per Saham (BVPS)', 'Rp 1.241,35'),
    ('Arus Kas Per Saham (CFPS)', 'Rp 15,77'),
    ('Kas Ekuivalen Per Saham (CEPS)', 'Rp 253,90'),
    ('Aset Bersih Per Saham (NAVS)', 'Rp 1.695,28'),
]

for label, value in per_share_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_lavender
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_lavender
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === VALUATION METRICS ===
cell = ws.cell(row=row, column=start_col, value='METRIK VALUASI')
cell.font = section_font
cell.fill = section_teal
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
valuation_data = [
    ('Imbal Hasil Dividen', '0,03%'),
    ('Rasio Harga/Laba (PER)', '216,79x'),
    ('Rasio Harga/Pendapatan (PSR)', '55,36x'),
    ('Rasio Harga/Nilai Buku (PBVR)', '10,17x'),
    ('Rasio Harga/Arus Kas (PCFR)', '800,79x'),
]

for label, value in valuation_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_sky
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_sky
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === PROFITABILITY ===
cell = ws.cell(row=row, column=start_col, value='RASIO PROFITABILITAS')
cell.font = section_font
cell.fill = section_pink
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
profitability_data = [
    ('Rasio Pembayaran Dividen (DPR)', '6,87%'),
    ('Margin Laba Kotor (GPM)', '65,61%'),
    ('Margin Laba Operasi (OPM)', '55,82%'),
    ('Margin Laba Bersih (NPM)', '25,54%'),
    ('Margin EBIT (EBITM)', '57,45%'),
    ('Imbal Hasil Ekuitas (ROE)', '4,69%'),
    ('Imbal Hasil Aset (ROA)', '2,13%'),
]

for label, value in profitability_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_rose
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_rose
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === LIQUIDITY ===
cell = ws.cell(row=row, column=start_col, value='RASIO LIKUIDITAS & SOLVABILITAS')
cell.font = section_font
cell.fill = section_navy
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
liquidity_data = [
    ('Rasio Utang/Ekuitas (DER)', '83,37%'),
    ('Rasio Kas (Cash Ratio)', '29,36%'),
    ('Rasio Cepat (Quick Ratio)', '47,14%'),
    ('Rasio Lancar (Current Ratio)', '194,82%'),
]

for label, value in liquidity_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_cream
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_cream
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === DIVIDEND HISTORY ===
cell = ws.cell(row=row, column=start_col, value='RIWAYAT DIVIDEN')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
dividend_data = [
    ('Dividen 2024', 'Rp 4/lembar (Ex: 26/05/2025)'),
    ('Dividen 2023', 'Rp 2/lembar (Ex: 05/07/2024)'),
]

for label, value in dividend_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_mint
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_mint
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === QUARTERLY PERFORMANCE ===
cell = ws.cell(row=row, column=start_col, value='KINERJA KUARTALAN 2025')
cell.font = section_font
cell.fill = section_orange
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
quarterly_data = [
    ('Pendapatan Q1-Q3 2025', 'Rp 4,13 Triliun'),
    ('Laba Bersih Q1-Q3 2025', 'Rp 1,06 Triliun'),
    ('EPS Kumulatif 2025', 'Rp 62'),
    ('Pertumbuhan YoY Revenue', '+45,8%'),
    ('Pertumbuhan YoY Net Income', '+69,1%'),
]

for label, value in quarterly_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_peach
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_peach
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

# Set column widths
ws.column_dimensions[get_column_letter(start_col)].width = 38
ws.column_dimensions[get_column_letter(start_col+1)].width = 32

# Save workbook
wb.save('C:/doc Herman/pani.xlsx')
print('PANI Fundamental data added to Excel successfully!')
print(f'Data written from column AL (38) to AM (39)')
print(f'Total rows used: {row}')
