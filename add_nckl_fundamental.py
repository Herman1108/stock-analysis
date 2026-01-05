"""
Script to add Fundamental data to NCKL Excel file
Starting from column AL with attractive, colorful design
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load workbook
wb = openpyxl.load_workbook('C:/doc Herman/nckl.xlsx')
ws = wb.active

# Define styles - Earthy and professional colors for mining company
header_fill = PatternFill(start_color='37474F', end_color='37474F', fill_type='solid')  # Dark slate
section_green = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')  # Forest green
section_orange = PatternFill(start_color='EF6C00', end_color='EF6C00', fill_type='solid')  # Deep orange
section_purple = PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid')  # Deep purple
section_teal = PatternFill(start_color='00838F', end_color='00838F', fill_type='solid')  # Dark cyan
section_pink = PatternFill(start_color='AD1457', end_color='AD1457', fill_type='solid')  # Dark pink
section_navy = PatternFill(start_color='283593', end_color='283593', fill_type='solid')  # Indigo
section_brown = PatternFill(start_color='5D4037', end_color='5D4037', fill_type='solid')  # Brown

data_mint = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')  # Light green
data_peach = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')  # Light orange
data_lavender = PatternFill(start_color='EDE7F6', end_color='EDE7F6', fill_type='solid')  # Light purple
data_sky = PatternFill(start_color='E0F7FA', end_color='E0F7FA', fill_type='solid')  # Light cyan
data_cream = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')  # Light amber
data_rose = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')  # Light pink
data_sand = PatternFill(start_color='EFEBE9', end_color='EFEBE9', fill_type='solid')  # Light brown

header_font = Font(bold=True, color='FFFFFF', size=12)
section_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, color='37474F', size=10)
normal_font = Font(color='37474F', size=10)
italic_font = Font(italic=True, color='78909C', size=9)
number_font = Font(color='1565C0', size=10)
negative_font = Font(color='C62828', size=10)

thin_border = Border(
    left=Side(style='thin', color='90A4AE'),
    right=Side(style='thin', color='90A4AE'),
    top=Side(style='thin', color='90A4AE'),
    bottom=Side(style='thin', color='90A4AE')
)

# Starting column AL = 38
start_col = 38
row = 1

# === MAIN HEADER ===
cell = ws.cell(row=row, column=start_col, value='ANALISIS FUNDAMENTAL')
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal='center')
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row = 3
# === MARKET OVERVIEW ===
cell = ws.cell(row=row, column=start_col, value='RINGKASAN PASAR')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Per 30 September 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
market_data = [
    ('Tahun Fiskal Berakhir', 'Desember'),
    ('Total Saham Beredar', '63,10 Miliar lembar'),
    ('Kapitalisasi Pasar', 'Rp 73,51 Triliun'),
    ('Indeks Saham', '93,2'),
]

for label, value in market_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_mint
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_mint
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === FINANCIAL POSITION ===
cell = ws.cell(row=row, column=start_col, value='POSISI KEUANGAN')
cell.font = section_font
cell.fill = section_orange
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
financial_data = [
    ('Pendapatan Usaha', 'Rp 22,40 Triliun'),
    ('Total Aset', 'Rp 58,54 Triliun'),
    ('Total Liabilitas', 'Rp 15,20 Triliun'),
    ('Total Ekuitas', 'Rp 35,72 Triliun'),
    ('Belanja Modal (CapEx)', 'Rp 451,84 Miliar'),
    ('Biaya Operasional', 'Rp 929,91 Miliar'),
    ('Arus Kas Operasional', 'Rp 7,28 Triliun'),
    ('Arus Kas Bersih', 'Rp -1,63 Triliun'),
    ('Laba Usaha', 'Rp 6,44 Triliun'),
    ('Laba Tahun Berjalan', 'Rp 6,45 Triliun'),
]

for label, value in financial_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_peach
    ws.cell(row=row, column=start_col).border = thin_border
    cell = ws.cell(row=row, column=start_col+1, value=value)
    cell.font = negative_font if '-' in value else number_font
    ws.cell(row=row, column=start_col+1).fill = data_peach
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === PER SHARE DATA ===
cell = ws.cell(row=row, column=start_col, value='DATA PER LEMBAR SAHAM')
cell.font = section_font
cell.fill = section_purple
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
per_share_data = [
    ('Dividen Per Saham (DPS)', 'Rp 30,36'),
    ('Laba Per Saham (EPS)', 'Rp 136,23'),
    ('Pendapatan Per Saham (RPS)', 'Rp 473,39'),
    ('Nilai Buku Per Saham (BVPS)', 'Rp 566,11'),
    ('Arus Kas Per Saham (CFPS)', 'Rp 153,81'),
    ('Kas Setara Per Saham (CEPS)', 'Rp 78,50'),
    ('Aset Bersih Per Saham (NAVS)', 'Rp 686,91'),
]

for label, value in per_share_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_lavender
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_lavender
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === VALUATION METRICS ===
cell = ws.cell(row=row, column=start_col, value='METRIK VALUASI')
cell.font = section_font
cell.fill = section_teal
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
valuation_data = [
    ('Yield Dividen', '2,61%'),
    ('Rasio Harga/Laba (PER)', '8,55x'),
    ('Rasio Harga/Pendapatan (PSR)', '2,46x'),
    ('Rasio Harga/Nilai Buku (PBV)', '2,06x'),
    ('Rasio Harga/Arus Kas (PCFR)', '7,57x'),
]

for label, value in valuation_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_sky
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_sky
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === PROFITABILITY ===
cell = ws.cell(row=row, column=start_col, value='RASIO PROFITABILITAS')
cell.font = section_font
cell.fill = section_pink
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
profitability_data = [
    ('Rasio Pembayaran Dividen (DPR)', '22,28%'),
    ('Marjin Laba Kotor (GPM)', '32,92%'),
    ('Marjin Laba Usaha (OPM)', '28,77%'),
    ('Marjin Laba Bersih (NPM)', '28,78%'),
    ('Marjin EBIT (EBITM)', '41,66%'),
    ('Imbal Hasil Ekuitas (ROE)', '24,07%'),
    ('Imbal Hasil Aset (ROA)', '14,68%'),
]

for label, value in profitability_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_rose
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_rose
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === LIQUIDITY ===
cell = ws.cell(row=row, column=start_col, value='RASIO LIKUIDITAS & LEVERAGE')
cell.font = section_font
cell.fill = section_navy
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
liquidity_data = [
    ('Rasio Utang/Ekuitas (DER)', '42,54%'),
    ('Rasio Kas (Cash Ratio)', '78,50%'),
    ('Rasio Cepat (Quick Ratio)', '112,16%'),
    ('Rasio Lancar (Current Ratio)', '195,71%'),
]

for label, value in liquidity_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_cream
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_cream
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === DIVIDEND HISTORY ===
cell = ws.cell(row=row, column=start_col, value='RIWAYAT DIVIDEN')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
dividend_data = [
    ('Dividen 2024', 'Rp 30,357 (Ex: 30/06/2025)'),
    ('Dividen 2023', 'Rp 26,716 (Ex: 08/07/2024)'),
    ('Dividen 2022', 'Rp 22,189 (Ex: 11/07/2023)'),
]

for label, value in dividend_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_mint
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_mint
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === QUARTERLY PERFORMANCE ===
cell = ws.cell(row=row, column=start_col, value='KINERJA KUARTALAN 2025')
cell.font = section_font
cell.fill = section_orange
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
quarterly_data = [
    ('Pendapatan Q1 2025', 'Rp 7,13 Triliun'),
    ('Pendapatan Q2 2025', 'Rp 6,97 Triliun'),
    ('Pendapatan Q3 2025', 'Rp 8,31 Triliun'),
    ('Total Pendapatan 9M 2025', 'Rp 22,40 Triliun'),
    ('Laba Bersih Q1 2025', 'Rp 1,66 Triliun'),
    ('Laba Bersih Q2 2025', 'Rp 2,45 Triliun'),
    ('Laba Bersih Q3 2025', 'Rp 2,34 Triliun'),
    ('Total Laba Bersih 9M 2025', 'Rp 6,45 Triliun'),
    ('EPS Kumulatif 2025', 'Rp 136'),
]

for label, value in quarterly_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_peach
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_peach
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === HISTORICAL COMPARISON ===
cell = ws.cell(row=row, column=start_col, value='PERBANDINGAN TAHUNAN')
cell.font = section_font
cell.fill = section_brown
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
historical_data = [
    ('EPS 2022', 'Rp 85'),
    ('EPS 2023', 'Rp 89'),
    ('EPS 2024', 'Rp 101'),
    ('EPS 2025 (9M)', 'Rp 136'),
    ('Pendapatan 2022', 'Rp 10 Triliun'),
    ('Pendapatan 2023', 'Rp 24 Triliun'),
    ('Pendapatan 2024', 'Rp 27 Triliun'),
    ('Laba Bersih 2022', 'Rp 5 Triliun'),
    ('Laba Bersih 2023', 'Rp 6 Triliun'),
    ('Laba Bersih 2024', 'Rp 6 Triliun'),
]

for label, value in historical_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_sand
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = number_font
    ws.cell(row=row, column=start_col+1).fill = data_sand
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

# Set column widths
ws.column_dimensions[get_column_letter(start_col)].width = 38
ws.column_dimensions[get_column_letter(start_col+1)].width = 32

# Save workbook
wb.save('C:/doc Herman/nckl.xlsx')
print('NCKL Fundamental data added to Excel successfully!')
print(f'Data written from column AL (38) to AM (39)')
print(f'Total rows used: {row}')
