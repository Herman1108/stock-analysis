"""
Script to add Company Profile data to CDIA Excel file
Starting from column AI with attractive, colorful design
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load workbook
wb = openpyxl.load_workbook('C:/doc Herman/cdia.xlsx')
ws = wb.active

# Define styles - Soft and cheerful colors
header_fill = PatternFill(start_color='4A90D9', end_color='4A90D9', fill_type='solid')  # Soft blue
section_green = PatternFill(start_color='7BC97F', end_color='7BC97F', fill_type='solid')  # Soft green
highlight_fill = PatternFill(start_color='F9E79F', end_color='F9E79F', fill_type='solid')  # Soft yellow
data_fill = PatternFill(start_color='E8F6F3', end_color='E8F6F3', fill_type='solid')  # Soft mint
pink_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')  # Soft pink
lavender_fill = PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid')  # Soft lavender
orange_fill = PatternFill(start_color='E59866', end_color='E59866', fill_type='solid')  # Soft orange
sky_blue = PatternFill(start_color='5DADE2', end_color='5DADE2', fill_type='solid')  # Sky blue
purple_fill = PatternFill(start_color='AF7AC5', end_color='AF7AC5', fill_type='solid')  # Soft purple
teal_fill = PatternFill(start_color='48C9B0', end_color='48C9B0', fill_type='solid')  # Teal
steel_blue = PatternFill(start_color='5499C7', end_color='5499C7', fill_type='solid')  # Steel blue
coral_fill = PatternFill(start_color='EC7063', end_color='EC7063', fill_type='solid')  # Soft coral

header_font = Font(bold=True, color='FFFFFF', size=12)
section_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, color='2C3E50', size=10)
normal_font = Font(color='2C3E50', size=10)
italic_font = Font(italic=True, color='566573', size=9)

thin_border = Border(
    left=Side(style='thin', color='BDC3C7'),
    right=Side(style='thin', color='BDC3C7'),
    top=Side(style='thin', color='BDC3C7'),
    bottom=Side(style='thin', color='BDC3C7')
)

# Starting column AI = 35
start_col = 35
row = 1

# === MAIN HEADER ===
cell = ws.cell(row=row, column=start_col, value='COMPANY PROFILE')
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal='center')
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row = 3
# === COMPANY IDENTITY ===
cell = ws.cell(row=row, column=start_col, value='IDENTITAS PERUSAHAAN')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
profile_data = [
    ('Nama Emiten', 'PT Chandra Daya Investasi Tbk'),
    ('Kode Saham', 'CDIA'),
    ('Papan Pencatatan', 'Papan Pengembangan (Development)'),
    ('Sektor Usaha', 'Infrastruktur'),
    ('Sub Sektor', 'Utilitas'),
    ('Bidang Industri', 'Utilitas Kelistrikan'),
    ('Aktivitas Bisnis', 'Holding company dan konsultasi manajemen investasi infrastruktur'),
]

for label, value in profile_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = data_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === COMPANY HISTORY ===
cell = ws.cell(row=row, column=start_col, value='SEJARAH & PENCATATAN')
cell.font = section_font
cell.fill = orange_fill
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
history_data = [
    ('Tanggal Pencatatan', '09 Juli 2025'),
    ('Tanggal Efektif', '30 Juni 2025'),
    ('Nilai Nominal', 'Rp 100'),
    ('Harga Penawaran Perdana', 'Rp 190'),
    ('Jumlah Saham IPO', '12.480.000.000 lembar'),
    ('Dana IPO Terhimpun', 'Rp 2,37 Triliun'),
    ('Penjamin Emisi', 'BCA, BNI, DBS Vickers, Henan Putihrai, OCBC, Trimegah Sekuritas'),
    ('Biro Administrasi Efek', 'PT Datindo Entrycom'),
]

for label, value in history_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = pink_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = pink_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === COMPANY BACKGROUND ===
cell = ws.cell(row=row, column=start_col, value='LATAR BELAKANG PERUSAHAAN')
cell.font = section_font
cell.fill = sky_blue
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
# Restructured company background (different wording from source)
background_text = [
    'Emiten ini fokus pada sektor investasi infrastruktur strategis Indonesia.',
    'Mendukung pertumbuhan ekonomi melalui proyek-proyek infrastruktur vital.',
    'Lini bisnis utama mencakup logistik dan transportasi maritim.',
    'Berperan dalam pengadaan armada kapal untuk distribusi logistik nasional.',
    'Tercatat di Bursa Efek Indonesia pada Juli 2025 dengan valuasi Rp 2,37 T.',
]

for text in background_text:
    ws.cell(row=row, column=start_col, value=text).font = normal_font
    ws.cell(row=row, column=start_col).fill = lavender_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)
    row += 1

row += 1
# === SHAREHOLDERS ===
cell = ws.cell(row=row, column=start_col, value='KOMPOSISI PEMEGANG SAHAM')
cell.font = section_font
cell.fill = purple_fill
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Per 30 November 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
shareholders = [
    ('PT Chandra Asri Pacific Tbk (Pengendali)', '60,00%'),
    ('Phoenix Power B.V.', '30,00%'),
    ('Publik (Non-Warkat)', '9,97%'),
    ('Total Saham Beredar', '124.829.374.700 lembar'),
]

for label, value in shareholders:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = highlight_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = highlight_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === SHAREHOLDER COUNT HISTORY ===
cell = ws.cell(row=row, column=start_col, value='PERKEMBANGAN JUMLAH INVESTOR')
cell.font = section_font
cell.fill = teal_fill
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
investor_history = [
    ('November 2025', '257.060 (-8.817)'),
    ('Oktober 2025', '265.877 (+54.866)'),
    ('September 2025', '211.011 (-51.275)'),
    ('Agustus 2025', '262.286 (-101.207)'),
    ('Juli 2025', '363.493 (-35.640)'),
    ('IPO 9 Juli 2025', '399.133 (+0)'),
]

for label, value in investor_history:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = data_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === BOARD OF DIRECTORS ===
cell = ws.cell(row=row, column=start_col, value='JAJARAN DIREKSI')
cell.font = section_font
cell.fill = steel_blue
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Efektif 9 Juli 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
directors = [
    ('Presiden Direktur', 'Fransiskus Ruly Aryawan'),
    ('Direktur', 'Jonathan Kandinata'),
    ('Direktur', 'Saksit Suntharekanon'),
    ('Direktur', 'Merly'),
    ('Direktur', 'Agus Lukmanul Hakim'),
]

for label, value in directors:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = pink_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = pink_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === BOARD OF COMMISSIONERS ===
cell = ws.cell(row=row, column=start_col, value='DEWAN KOMISARIS')
cell.font = section_font
cell.fill = coral_fill
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Efektif 9 Juli 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
commissioners = [
    ('Presiden Komisaris Independen', 'Erry Riyana Hardjapamekas'),
    ('Komisaris Independen', 'Ade Supandi, SE'),
    ('Komisaris', 'Erwin Ciputra'),
    ('Komisaris', 'Andre Khor Kah Hin'),
    ('Komisaris', 'Thawat Hirancharukorn'),
    ('Komisaris', 'Prasit Laohawirapap'),
]

for label, value in commissioners:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = lavender_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = lavender_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

# Set column widths
ws.column_dimensions[get_column_letter(start_col)].width = 35
ws.column_dimensions[get_column_letter(start_col+1)].width = 45

# Save workbook
wb.save('C:/doc Herman/cdia.xlsx')
print('CDIA Profile data added to Excel successfully!')
print(f'Data written from column AI (35) to AJ (36)')
print(f'Total rows used: {row}')
