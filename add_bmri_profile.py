"""
Script to add Company Profile data to BMRI Excel file
Starting from column AI with attractive, colorful design (same format as CDIA)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load workbook
wb = openpyxl.load_workbook('C:/doc Herman/bmri.xlsx')
ws = wb.active

# Define styles - Soft and cheerful colors (same as CDIA)
header_fill = PatternFill(start_color='4A90D9', end_color='4A90D9', fill_type='solid')  # Soft blue
section_green = PatternFill(start_color='7BC97F', end_color='7BC97F', fill_type='solid')  # Soft green
highlight_fill = PatternFill(start_color='F9E79F', end_color='F9E79F', fill_type='solid')  # Soft yellow
data_fill = PatternFill(start_color='E8F6F3', end_color='E8F6F3', fill_type='solid')  # Soft mint
pink_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')  # Soft pink
lavender_fill = PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid')  # Soft lavender
orange_fill = PatternFill(start_color='E59866', end_color='E59866', fill_type='solid')  # Soft orange
sky_blue = PatternFill(start_color='5DADE2', end_color='5DADE2', fill_type='solid')  # Sky blue
purple_fill = PatternFill(start_color='AF7AC5', end_color='AF7AC5', fill_type='solid')  # Soft purple
teal_fill = PatternFill(start_color='48C9B0', end_color='48C9B0', fill_type='solid')  # Teal
steel_blue = PatternFill(start_color='5499C7', end_color='5499C7', fill_type='solid')  # Steel blue
coral_fill = PatternFill(start_color='EC7063', end_color='EC7063', fill_type='solid')  # Soft coral

header_font = Font(bold=True, color='FFFFFF', size=12)
section_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, color='2C3E50', size=10)
normal_font = Font(color='2C3E50', size=10)
italic_font = Font(italic=True, color='566573', size=9)

thin_border = Border(
    left=Side(style='thin', color='BDC3C7'),
    right=Side(style='thin', color='BDC3C7'),
    top=Side(style='thin', color='BDC3C7'),
    bottom=Side(style='thin', color='BDC3C7')
)

# Starting column AI = 35
start_col = 35
row = 1

# === MAIN HEADER ===
cell = ws.cell(row=row, column=start_col, value='COMPANY PROFILE')
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal='center')
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row = 3
# === COMPANY IDENTITY ===
cell = ws.cell(row=row, column=start_col, value='IDENTITAS PERUSAHAAN')
cell.font = section_font
cell.fill = section_green
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
profile_data = [
    ('Nama Emiten', 'PT Bank Mandiri (Persero) Tbk'),
    ('Kode Saham', 'BMRI'),
    ('Papan Pencatatan', 'Papan Utama (Main Board)'),
    ('Sektor Usaha', 'Keuangan (Financials)'),
    ('Sub Sektor', 'Perbankan'),
    ('Bidang Industri', 'Bank'),
    ('Aktivitas Bisnis', 'Jasa Keuangan - Perbankan'),
]

for label, value in profile_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = data_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === COMPANY HISTORY ===
cell = ws.cell(row=row, column=start_col, value='SEJARAH & PENCATATAN')
cell.font = section_font
cell.fill = orange_fill
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
history_data = [
    ('Tanggal Pencatatan', '14 Juli 2003'),
    ('Tanggal Efektif', '27 Juni 2003'),
    ('Nilai Nominal', 'Rp 125'),
    ('Harga Penawaran Perdana', 'Rp 675'),
    ('Jumlah Saham IPO', '2,90 Miliar lembar'),
    ('Dana IPO Terhimpun', 'Rp 1,96 Triliun'),
    ('Penjamin Emisi', 'PT Danareksa Sekuritas, PT ABN Amro Asia Securities'),
    ('Biro Administrasi Efek', 'PT Datindo Entrycom'),
]

for label, value in history_data:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = pink_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = pink_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === COMPANY BACKGROUND ===
cell = ws.cell(row=row, column=start_col, value='LATAR BELAKANG PERUSAHAAN')
cell.font = section_font
cell.fill = sky_blue
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
background_text = [
    'Bank BUMN terbesar di Indonesia dengan layanan perbankan komprehensif.',
    'Anak usaha: Bank Syariah Mandiri, Mandiri Sekuritas, Bank Sinar Harapan Bali.',
    'Mandiri Tunas Finance (pembiayaan), Mandiri Internasional Remittance.',
    'AXA Mandiri Financial Service (asuransi jiwa), Mandiri AXA General Insurance.',
    'Bank Mandiri (Europe) Limited untuk ekspansi perbankan internasional.',
]

for text in background_text:
    ws.cell(row=row, column=start_col, value=text).font = normal_font
    ws.cell(row=row, column=start_col).fill = lavender_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)
    row += 1

row += 1
# === SHAREHOLDERS ===
cell = ws.cell(row=row, column=start_col, value='KOMPOSISI PEMEGANG SAHAM')
cell.font = section_font
cell.fill = purple_fill
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Per 30 November 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
shareholders = [
    ('PT Danantara Asset Management (Persero)', '52,00%'),
    ('Masyarakat Non Warkat (Scripless)', '39,86%'),
    ('Indonesia Investment Authority', '8,00%'),
    ('Saham Treasury', '0,0753%'),
    ('Negara Republik Indonesia', '<0,0001%'),
    ('Total Saham Beredar', '93.333.333.332 lembar'),
]

for label, value in shareholders:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = highlight_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = highlight_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === SHAREHOLDER COUNT HISTORY ===
cell = ws.cell(row=row, column=start_col, value='PERKEMBANGAN JUMLAH INVESTOR')
cell.font = section_font
cell.fill = teal_fill
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
investor_history = [
    ('November 2025', '279.942 (-30.914)'),
    ('Oktober 2025', '310.856 (-3.370)'),
    ('September 2025', '314.226 (+42.263)'),
    ('Agustus 2025', '271.963 (+15.707)'),
    ('Juli 2025', '287.632 (+31.376)'),
    ('Juni 2025', '256.256 (+5.007)'),
]

for label, value in investor_history:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = data_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = data_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === BOARD OF DIRECTORS ===
cell = ws.cell(row=row, column=start_col, value='JAJARAN DIREKSI')
cell.font = section_font
cell.fill = steel_blue
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Efektif 23 Desember 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
directors = [
    ('Direktur Utama', 'Riduan'),
    ('Wakil Direktur Utama', 'Henry Panjaitan'),
    ('Direktur', 'Timothy Utama'),
    ('Direktur', 'Eka Fitria'),
    ('Direktur', 'Danis Subyantoro'),
    ('Direktur', 'Totok Priyambodo'),
    ('Direktur', 'Mochamad Rizaldi'),
    ('Direktur', 'Saptari'),
    ('Direktur', 'Ari Rizaldi'),
    ('Direktur', 'Novita Widya Anggraeni'),
    ('Direktur', 'Sunarto'),
    ('Direktur', 'Jan Winston Tambunan'),
]

for label, value in directors:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = pink_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = pink_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

row += 1
# === BOARD OF COMMISSIONERS ===
cell = ws.cell(row=row, column=start_col, value='DEWAN KOMISARIS')
cell.font = section_font
cell.fill = coral_fill
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
ws.cell(row=row, column=start_col, value='Efektif 23 Desember 2025').font = italic_font
ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)

row += 1
commissioners = [
    ('Komisaris Utama', 'Zulkifli Zaini (Independen)'),
    ('Wakil Komisaris Utama', 'M. Rudy Salahuddin Ramto'),
    ('Komisaris', 'Luky Alfirman'),
    ('Komisaris', 'Muhammad Yusuf Ateh'),
    ('Komisaris', 'Yuliot'),
    ('Komisaris Independen', 'B. Bintoro Kunto Pardewo'),
    ('Komisaris Independen', 'Mia Amiati'),
]

for label, value in commissioners:
    ws.cell(row=row, column=start_col, value=label).font = title_font
    ws.cell(row=row, column=start_col).fill = lavender_fill
    ws.cell(row=row, column=start_col).border = thin_border
    ws.cell(row=row, column=start_col+1, value=value).font = normal_font
    ws.cell(row=row, column=start_col+1).fill = lavender_fill
    ws.cell(row=row, column=start_col+1).border = thin_border
    row += 1

# Set column widths (same as CDIA)
ws.column_dimensions[get_column_letter(start_col)].width = 35
ws.column_dimensions[get_column_letter(start_col+1)].width = 45

# Save workbook
wb.save('C:/doc Herman/bmri.xlsx')
print('BMRI Profile data added to Excel successfully!')
print(f'Data written from column AI (35) to AJ (36)')
print(f'Total rows used: {row}')
